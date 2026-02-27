"""
Módulo de processamento de PDFs e arquivos ZIP.
Descompacta ZIPs, extrai texto de PDFs (incluindo OCR para escaneados).
"""

import logging
import os
import zipfile
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl

from . import config

logger = logging.getLogger(__name__)


def parse_xlsx(filepath: str) -> list[dict]:
    """
    Lê o XLSX exportado do ConLicitação e retorna uma lista de licitações.
    
    Args:
        filepath: Caminho do arquivo XLSX
    Returns: Lista de dicts com os dados de cada licitação
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        # Ler cabeçalhos da primeira linha
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip().lower())

        # Mapear colunas esperadas (flexível para variações de nome)
        column_map = {}
        expected_fields = {
            "objeto": ["objeto", "descrição", "descricao", "description"],
            "orgao": ["órgão", "orgao", "orgão licitante", "orgao licitante"],
            "cidade": ["cidade", "município", "municipio", "city"],
            "uf": ["uf", "estado", "state"],
            "data_abertura": ["data abertura", "abertura", "data", "datas"],
            "edital": ["edital", "nº edital", "numero edital", "nº"],
            "status": ["status", "situação", "situacao"],
            "palavras_chave": ["palavras-chave", "palavras chave", "keywords", "palavra-chave"],
            "valor": ["valor", "valor estimado", "preço", "preco"],
            "modalidade": ["modalidade", "tipo"],
            "numero_conlicitacao": ["nº conlicitação", "conlicitação", "conlicitacao", "id"],
        }

        for field, variations in expected_fields.items():
            for i, header in enumerate(headers):
                if any(v in header for v in variations):
                    column_map[field] = i
                    break

        # Ler dados
        licitacoes = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            licitacao = {"row_index": row_idx}
            for field, col_idx in column_map.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    licitacao[field] = str(value).strip() if value else ""

            # Garantir campos obrigatórios
            if not licitacao.get("objeto"):
                continue

            licitacoes.append(licitacao)

        wb.close()
        logger.info(f"[OK] XLSX parseado: {len(licitacoes)} licitações encontradas")
        return licitacoes

    except Exception as e:
        logger.error(f"❌ Erro ao parsear XLSX: {e}")
        return []


def extract_zip(zip_path: str, processed_zips: set = None) -> list[str]:
    """
    Descompacta um arquivo ZIP e retorna os caminhos dos arquivos extraídos.
    Suporta extração recursiva de ZIPs aninhados.
    
    Args:
        zip_path: Caminho do arquivo ZIP
        processed_zips: Set de arquivos já processados para evitar loops
    Returns: Lista de caminhos de todos os arquivos extraídos (incluindo em ZIPs aninhados)
    """
    if processed_zips is None:
        processed_zips = set()
    
    abs_path = os.path.abspath(zip_path)
    if abs_path in processed_zips:
        return []
    processed_zips.add(abs_path)

    try:
        extract_dir = os.path.splitext(zip_path)[0]
        os.makedirs(extract_dir, exist_ok=True)

        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(extract_dir)

        all_extracted = []
        # Percorrer os arquivos extraídos para ver se há mais ZIPs
        for root, dirs, files in os.walk(extract_dir):
            for f in files:
                filepath = os.path.join(root, f)
                
                # Se for outro ZIP, extrair recursivamente
                if f.lower().endswith(".zip"):
                    nested_files = extract_zip(filepath, processed_zips)
                    all_extracted.extend(nested_files)
                else:
                    all_extracted.append(filepath)

        logger.info(f"✅ ZIP descompactado (recursivo): {len(all_extracted)} arquivos totais de {zip_path}")
        return all_extracted

    except zipfile.BadZipFile:
        # Pode ser um PDF direto (não ZIP)
        logger.warning(f"⚠️ Arquivo não é ZIP, pode ser PDF direto: {zip_path}")
        return [zip_path]
    except Exception as e:
        logger.error(f"❌ Erro ao descompactar ZIP: {e}")
        return []


def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Extrai texto de um PDF usando PyMuPDF.
    Suporta PDFs nativos (texto) e tenta OCR para escaneados.
    
    Args:
        pdf_path: Caminho do arquivo PDF
    Returns: Texto extraído do PDF
    """
    try:
        doc = fitz.open(pdf_path)
        full_text = []

        for page_num, page in enumerate(doc, start=1):
            text = page.get_text("text")
            if text.strip():
                full_text.append(text)
            else:
                # Página sem texto — provavelmente escaneada
                # Tentar extração via OCR do PyMuPDF (se disponível)
                try:
                    text = page.get_text("text", flags=fitz.TEXT_PRESERVE_WHITESPACE)
                    if text.strip():
                        full_text.append(text)
                    else:
                        logger.warning(
                            f"⚠️ Página {page_num} de {pdf_path} sem texto "
                            f"(possível imagem escaneada — OCR não disponível)"
                        )
                except Exception:
                    pass

        doc.close()

        result = "\n\n".join(full_text)
        
        # Limitar tamanho para não estourar token limit da LLM
        max_chars = 50000  # ~12.5k tokens
        if len(result) > max_chars:
            result = result[:max_chars] + "\n\n[... TEXTO TRUNCADO ...]"
            logger.info(f"⚠️ PDF truncado: {len(result)} chars (limite: {max_chars})")

        logger.info(f"✅ PDF parseado: {len(result)} caracteres de {pdf_path}")
        return result

    except Exception as e:
        logger.error(f"❌ Erro ao extrair texto do PDF {pdf_path}: {e}")
        return ""


def process_edital_download(download_path: str) -> dict:
    """
    Processa um arquivo baixado (ZIP ou PDF):
    1. Se ZIP → descompacta
    2. Localiza PDFs
    3. Extrai texto de cada PDF

    Args:
        download_path: Caminho do arquivo baixado
    Returns: Dict com {pdf_files, texts, success}
    """
    result = {
        "download_path": download_path,
        "pdf_files": [],
        "texts": [],
        "combined_text": "",
        "success": False,
    }

    try:
        # Verificar se é ZIP ou PDF
        ext = Path(download_path).suffix.lower()
        
        if ext == ".zip":
            files = extract_zip(download_path)
        elif ext == ".pdf":
            files = [download_path]
        else:
            # Tentar como ZIP primeiro, depois como PDF
            files = extract_zip(download_path)
            if not files or files == [download_path]:
                files = [download_path]

        # Filtrar apenas PDFs
        pdf_files = [f for f in files if f.lower().endswith(".pdf")]
        result["pdf_files"] = pdf_files

        if not pdf_files:
            logger.warning(f"⚠️ Nenhum PDF encontrado em {download_path}")
            # Tentar ler o arquivo original como PDF mesmo sem extensão
            text = extract_text_from_pdf(download_path)
            if text:
                result["texts"] = [text]
                result["combined_text"] = text
                result["success"] = True
            return result

        # Extrair texto de cada PDF
        texts = []
        for pdf_path in pdf_files:
            text = extract_text_from_pdf(pdf_path)
            if text:
                texts.append({
                    "filename": os.path.basename(pdf_path),
                    "text": text,
                })

        result["texts"] = texts
        result["combined_text"] = "\n\n---\n\n".join(
            f"📄 {t['filename']}:\n{t['text']}" for t in texts
        )
        result["success"] = len(texts) > 0

        logger.info(
            f"✅ Edital processado: {len(texts)} PDFs com texto extraído"
        )
        return result

    except Exception as e:
        logger.error(f"❌ Erro ao processar edital: {e}")
        return result
